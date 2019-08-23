import openpyxl
import datetime

from openpyxl.styles import Color, PatternFill, Font, Border

wb = openpyxl.load_workbook("Test12.xlsx")
a = wb.active

b = a.max_row
c = a.max_column

cc = 0  # cc means completed count
ip = 0  # ip means in progress count
incp = 0 # incp means incomplete count
count_d = 0 # count is the count of all the milestones having deadline before today
count_ad = 0 # count_ad is the count of milestones having deadline after today
list_incp_d = []
list_ip = []
list_cc = []
list_incp = []
list_activities = []

for i in range(2,b+1,1):
    if a.cell(row = i, column = 1).value != None :
        list_activities.append(i)

for ii in range(2,b+1,1):
    while a.cell(row = ii, column = 3).value <= datetime.datetime.today(): # d is the column having values as dates
        count_d += 1
        if a.cell(row = ii, column = 5).value == None :  # e is the column having status
            incp += 1
            list_incp.append(ii)
        elif a.cell(row = ii, column = 5).value.upper() == "COMPLETED" :  # e is the column having status
            cc += 1
            list_cc.append(ii)
        else :
            incp += 1
            list_incp.append(ii)

    while a.cell(row=ii, column=3).value > datetime.datetime.today():
        count_ad += 1
        if a.cell(row = ii, column = 5).value == None :  # e is the column having status
            ip += 1
            list_ip.append(ii)
        elif a.cell(row = ii, column = 5).value.upper() == "COMPLETED" :  # e is the column having status
            list_cc.append(ii)
        elif a.cell(row = ii, column = 5).value.upper() == "IN PROGRESS" :
            ip += 1
            list_ip.append(ii)
        else :
            ip += 1
            list_ip.append(ii)


wb.create_sheet(title = "Report")
bb = wb.get_sheet_by_name("Report")

bb.column_dimensions['A'].width = 30
bb.column_dimensions['B'].width = 30
bb.column_dimensions['C'].width = 30
bb.column_dimensions['D'].width = 30
bb.column_dimensions['E'].width = 30
bb.column_dimensions['F'].width = 30
bb.column_dimensions['G'].width = 30

bb.cell(row = 2,column = 1).value = "Overall Brand Indicator"
bb.cell(row = 2, column = 3).value = str(round(100*cc/count_d)) + "%" + " of the project is on time"
bb.cell(row = 5,column = 1).value = "Activities"
bb.cell(row = 5,column = 2).value = "Milestones"
bb.cell(row = 5,column = 3).value = "Deadline"
bb.cell(row = 5,column = 4).value = "Person in Charge"
bb.cell(row = 5,column = 5).value = "Status"
bb.cell(row = 5,column = 6).value = "Date of Completion"
bb.cell(row = 5,column = 7).value = "Comments"

greyFill = PatternFill(start_color = 'FFC8C8C8', end_color = 'FFC8C8C8', fill_type = 'solid')
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
greenFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
yellowFill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

for q in range(1,c+1,1):
    bb.cell(row = 5, column = q).fill = greyFill

y = cc/count_d
if y > 0.85 :
    bb.cell(row = 2, column = 2).fill = greenFill
elif y > 0.7 :
    bb.cell(row=2, column=2).fill = yellowFill
else :
    bb.cell(row=2, column=2).fill = redFill

for qq in range(1,c+1,1):
    bb.cell(row = 7, column = qq).fill = redFill

bb.cell(row = 7,column = 1).value = "Details of Incomplete Tasks"

for j in range(0, len(list_incp),1) :
    for k in range(0, c, 1) :
         bb.cell(row = j+8, column = k + 2 ).value = a.cell(row = int(list_incp[j]), column = k+2).value
    for v in range(0, len(list_activities) - 1,1):
        if int(list_incp[j]) in list_activities :
                 bb.cell(row=j + 8, column= 1).value = a.cell(row=int(list_incp[j]), column= 1).value
                 break
        elif list_activities[v] < int(list_incp[j]) and int(list_incp[j] ) < list_activities[v+1] :
                 bb.cell(row=j + 8, column=1).value = a.cell(row= list_activities[v] , column=1).value
                 break
        elif int(list_incp[j]) > list_activities[len(list_activities) - 1] :
                 bb.cell(row=j + 8, column=1).value = a.cell(row=list_activities[len(list_activities) - 1], column=1).value
        else :
                 continue

for qqq in range(1,c+1,1):
    bb.cell(row = 4 + len(list_incp) + 6, column = qqq).fill = yellowFill

bb.cell(row = 4 + len(list_incp) + 6,column = 1).value = "Details of tasks which are in progress"

for jj in range(0, len(list_ip),1) :
    for kk in range(0, c, 1) :
         bb.cell(row = jj+4+len(list_incp)+7, column = kk + 2 ).value = a.cell(row = int(list_ip[jj]), column = kk+2).value
    for vv in range(0, len(list_activities)-1,1):
        if int(list_ip[jj]) in list_activities :
                 bb.cell(row= jj+4+len(list_incp)+7, column= 1).value = a.cell(row=int(list_ip[jj]), column= 1).value
                 break
        elif list_activities[vv] < int(list_ip[jj]) and int(list_ip[jj] ) < list_activities[vv+1] :
                 bb.cell(row=jj+4+len(list_incp)+7, column=1).value = a.cell(row= list_activities[vv] , column=1).value
                 break
        elif int(list_ip[jj]) > list_activities[len(list_activities) - 1] :
                 bb.cell(row=jj+4+len(list_incp)+7, column=1).value = a.cell(row=list_activities[len(list_activities) - 1], column=1).value
        else :
                 continue

for qqqq in range(1,c+1,1):
    bb.cell(row = 4+len(list_incp)+3+len(list_ip)+6 , column = qqqq).fill = greenFill

bb.cell(row = 4+len(list_incp)+3+len(list_ip)+6,column = 1).value = "Details of Completed Tasks"

for jjj in range(0, len(list_cc),1) :
    for kkk in range(0, c, 1) :
         bb.cell(row = jjj+4+len(list_incp)+3+len(list_ip)+7, column = kkk + 2 ).value = a.cell(row = int(list_cc[jjj]), column = kkk+2).value
    for vvv in range(0, len(list_activities) -1 ,1):
        if int(list_cc[jjj]) in list_activities :
                 bb.cell(row= jjj+4+len(list_incp)+3+len(list_ip)+7, column= 1).value = a.cell(row=int(list_cc[jjj]), column= 1).value
                 break
        elif list_activities[vvv] < int(list_cc[jjj]) and int(list_cc[jjj] ) < list_activities[vvv+1] :
                 bb.cell(row=jjj+4+len(list_incp)+3+len(list_ip)+7, column=1).value = a.cell(row= list_activities[vvv] , column=1).value
                 break
        elif int(list_cc[jjj]) > int(list_activities[len(list_activities) - 1]):
            bb.cell(row=jjj + 4 + len(list_incp) + 3 + len(list_ip) + 7, column=1).value = a.cell(row=list_activities[len(list_activities) - 1], column=1).value
        else :
                 continue

wb.save("Test12.xlsx")


